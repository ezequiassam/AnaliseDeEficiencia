import random
import timeit
import tracemalloc
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.setrecursionlimit(9000)


class MainFunctions:
    COUNT = 0
    random_vector_old = []

    def _swap(self, arr, index_i, index_j):
        arr[index_i], arr[index_j] = arr[index_j], arr[index_i]

    def count(self):
        self.COUNT = self.COUNT + 1

    def clear_count(self):
        self.COUNT = 0

    def average_of_limit(self, limit):
        average = self.COUNT / limit
        return str(float('%g' % average)).replace(".", ",")

    def exec_function(self, random_vector, element):
        pass

    def randint_vector(self, start, end):
        random_vector = [random.randint(start, end) for i in range(end)]
        if self.random_vector_old == random_vector:
            print("Vetor aleatório igual ao anterior")
            return self.randint_vector(start, end, random_vector)
        self.random_vector_old = random_vector
        return random_vector

    def repeat_average(self, limit, start=0, end=1000):
        self.clear_count()
        for iter in range(limit):
            random_vector = self.randint_vector(start, end)
            search_element = random.randint(start, end)
            self.exec_function(random_vector, search_element)

        average_str = self.average_of_limit(limit)
        return average_str

    def print_format_table(self, dict_csv):
        dict_format = {"1": [], "10": [], "100": [], "1000": [], "5000": []}
        for key, value in dict_csv.items():
            dict_format["1"].append(value[0])
            dict_format["10"].append(value[1])
            dict_format["100"].append(value[2])
            dict_format["1000"].append(value[3])
            dict_format["5000"].append(value[4])

        print(f"\n{'-----' * 45}\n\n")
        print(
            f"{'LOOP':<10}\t{'EXP. 1':<10}\t{'EXP. 2':<10}\t{'EXP. 3':<10}\t{'EXP. 4':<10}\t{'EXP. 5':<10}\t{'EXP. 6':<10}\t{'EXP. 7':<10}\t{'EXP. 8':<10}\t{'EXP. 9':<10}\t{'EXP. 10':<10}\t{'MEM PEAK':<10}\t{'TOTAL TEMP':<10}\t{'MEM CURRENT':<10}"
        )
        print(
            f"{'-----':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'------':<10}\t{'-------':<10}\t{'------':<10}\t{'-------':<10}\t{'------':<10}"
        )
        for loop, value in dict_format.items():
            print(
                "{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}|\t{:<10}"
                .format(loop, *value))
        return dict_format

    def print_format_table_end(self, start_function):
        peak, current = self._current_memory()
        temp = self._stop_function(start_function)
        print("{:<170}|\t{:<10}|\t{:<10}|\t{:<10}".format(
            "", f"{peak}MB", f"{temp}ms", f"{current}MB"))
        print(f"{'-----' * 40}\n\n")
        return peak, temp, current

    def save_table(self, dict_format, input_max, peak, temp, current):
        def get_list(index):
            if index == 1:
                return dict_format["1"]
            elif index == 2:
                return dict_format["10"]
            elif index == 3:
                return dict_format["100"]
            elif index == 4:
                return dict_format["1000"]
            elif index == 5:
                return dict_format["5000"]

        # Cria Novo workbook
        wb = Workbook()
        # Seleciona a active Sheet
        ws1 = wb.active
        # Rename it
        title = f"{str(type(self).__name__).upper()}-VALOR_DE_ENTRADA_{input_max}"
        ws1.title = title
        # Escreve alguns dados
        for col in range(1, 15):
            letter = get_column_letter(col)
            value = "LOOP"
            if col == 12:
                value = "MEM PEAK"
            if col == 13:
                value = "TOTAL TEMP"
            if col == 14:
                value = "MEM CURRENT"
            if col != 1 and col <= 11:
                value = f"EXP. {col - 1}"
            ws1[letter + str(1)] = value

        for col in range(1, 15):
            for row in range(2, 7):
                letter = get_column_letter(col)
                row_value = ""
                if col == 1:
                    row_value = str(list(dict_format.keys())[row - 2])
                elif col == 12 and row == 2:
                    row_value = f"{peak}MB"
                elif col == 13 and row == 2:
                    row_value = f"{temp}ms"
                elif col == 14 and row == 2:
                    row_value = f"{current}MB"
                elif col >= 2 and col <= 11:
                    row_value = get_list(row - 1)[col - 2]

                ws1[letter + str(row)] = row_value
        # Salva arquivo (Se não colocar o caminho complete, ele salva
        # na mesma pasta do scritp.
        wb.save(f"{title}.xlsx")

    def experiment_manager(self, input_max):
        dict_csv = {}
        print(
            f"EXECUTANDO FUNÇÃO {str(type(self).__name__).upper()} COM VALOR DE ENTRADA = {input_max}"
        )
        for iter in range(10):
            print(f"******** EXPERIMENTO {iter + 1} ********", end="\r")
            loop_1 = self.repeat_average(1, end=input_max)
            loop_10 = self.repeat_average(10, end=input_max)
            loop_100 = self.repeat_average(100, end=input_max)
            loop_1000 = self.repeat_average(1000, end=input_max)
            loop_5000 = self.repeat_average(5000, end=input_max)

            dict_csv[iter] = [loop_1, loop_10, loop_100, loop_1000, loop_5000]
        return self.print_format_table(dict_csv)

    def main(self, input_max=1000):
        start_function = self._star_function()
        dict_format = self.experiment_manager(input_max)
        peak, temp, current = self.print_format_table_end(start_function)
        self.save_table(dict_format, input_max, peak, temp, current)

    def _star_function(self):
        start_function = timeit.default_timer()
        tracemalloc.start()
        return start_function

    def _stop_function(self, start_function):
        tracemalloc.stop()
        end_function = timeit.default_timer()
        return (end_function - start_function)

    def _current_memory(self):
        current, peak = tracemalloc.get_traced_memory()
        return (current / 10**6), (peak / 10**6)


class BubbleSort(MainFunctions):
    def exec_function(self, random_vector, item):
        self.bubble_sort(random_vector)

    def bubble_sort(self, arr):
        self.count()
        for i in range(len(arr) - 1):
            self.count()
            swapped = False
            for j in range(len(arr) - 1):
                self.count()
                if arr[j] > arr[j + 1]:
                    self.count()
                    self._swap(arr, j, j + 1)
                    swapped = True
            if not swapped:
                self.count()
                break
        return arr


class MergeSort(MainFunctions):
    def exec_function(self, random_vector, item):
        self.mergesort(random_vector)

    def merge(self, rigth, left):
        arr_aux = []
        while len(rigth) != 0 and len(left) != 0:
            self.count()
            if rigth[0] > left[0]:
                self.count()
                arr_aux.append(left[0])
                del left[0]
            else:
                self.count()
                arr_aux.append(rigth[0])
                del rigth[0]

        while len(rigth) != 0:
            self.count()
            arr_aux.append(rigth[0])
            del rigth[0]

        while len(left) != 0:
            self.count()
            arr_aux.append(left[0])
            del left[0]

        return arr_aux

    def mergesort(self, arr):
        if len(arr) == 1:
            self.count()
            return arr

        middle = len(arr) // 2
        rigth = arr[0:middle]
        left = arr[middle:len(arr)]

        rigth = self.mergesort(rigth)
        left = self.mergesort(left)
        self.count()
        return self.merge(rigth, left)


class QuickSort(MainFunctions):
    def exec_function(self, random_vector, item):
        self.quick_sort(random_vector, 0, len(random_vector) - 1)

    def partition(self, arr, low, high):
        self.count()
        i = (low - 1)
        pivot = arr[high]  #ultimo elemento
        # pivot = arr[low] #primeiro elemento
        # pivot = arr[(low+high)//2] #elemento do meio
        # pivot = arr[random.randint(low, high)]  #elemento aleatorio

        for j in range(low, high):
            self.count()

            if arr[j] <= pivot:
                self.count()
                i = i + 1
                self._swap(arr, i, j)

        new_pi = i + 1
        self._swap(arr, new_pi, high)
        return new_pi

    def quick_sort(self, arr, low, high):
        self.count()
        if low < high:
            self.count()
            pi = self.partition(arr, low, high)

            self.quick_sort(arr, low, pi - 1)
            self.quick_sort(arr, pi + 1, high)


class InsertSort(MainFunctions):
    def exec_function(self, random_vector, item):
        self.insertion_sort(random_vector)

    def insertion_sort(self, arr):
        self.count()
        for i in range(1, len(arr)):
            self.count()
            key = arr[i]
            j = i - 1
            while j >= 0 and key < arr[j]:
                self.count()
                arr[j + 1] = arr[j]
                j -= 1
            arr[j + 1] = key


if __name__ == '__main__':

    def init_script(script_type: MainFunctions, input_max, auto):
        if not auto:
            script_type.main(input_max)
        else:
            for iter in [10, 100, 1000, 5000]:
                script_type.main(iter)

    init_script(QuickSort(), 10, False)
    # init_script(InsertSort(), 5000, False)
    #init_script(BubbleSort(), 5000, False)
    # init_script(InsertSort(), 0, True)
    # init_script(BubbleSort(), 0, True)
    # init_script(MergeSort(), 0, True)
    # init_script(QuickSort(), 0, True)
    """
    init_script(BubbleSort(), 1000, False)
    init_script(InsertSort(), 5000, False)
    init_script(BubbleSort(), 5000, False)

    loop = True
    while loop:
        input_max = 0
        input_auto = False
        cmd = int(
            input(
                "Digite o respectivo número para iniciar\n0 - Insertsort    1 - Bubblesort    2 - Mergesort    3 - Quicksort    4 - Sair\n"
            ))
        cmd_str = str(
            input(
                "Executa de forma automatica com valor de entrada pré definido? Digite S para sim ou N para não\n"
            ))
        if "S" in str(cmd_str).upper():
            input_auto = True
        else:
            input_max = int(input("Digite o Valor de Entrada\n"))

        if cmd == 0:
            init_script(InsertSort(), input_max, input_auto)
            loop = False
        elif cmd == 1:
            init_script(BubbleSort(), input_max, input_auto)
            loop = False
        elif cmd == 2:
            init_script(MergeSort(), input_max, input_auto)
            loop = False
        elif cmd == 3:
            init_script(QuickSort(), input_max, input_auto)
            loop = False
        elif cmd == 4:
            loop = False
        else:
            print("Informe os parâmetros corretamente")
            continue
"""
